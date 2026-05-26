import time
import json
import uuid
import zipfile
import io
import os
import pandas as pd
import openpyxl
import requests
from datetime import datetime, date, timedelta, timezone
from io import BytesIO

from utils.logger import log_event


class OrderFeedPrivateAPI:
    """Клиент для приватного API ленты заказов Wildberries"""
    def __init__(self, authorize_v3: str, wb_seller_lk: str, cookie: str):
        self.authorize_v3 = authorize_v3
        self.wb_seller_lk = wb_seller_lk
        self.cookie = cookie
        self.base_headers = {
            "Accept": "*/*",
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
            "Cookie": self.cookie,
            "Origin": "https://seller.wildberries.ru",
            "Referer": "https://seller.wildberries.ru/",
            "sec-ch-ua": '"Google Chrome";v="147", "Not.A/Brand";v="8", "Chromium";v="147"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Windows"',
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
        }

    def set_period(self, start_date: str, end_date: str):
        """Устанавливает период в интерфейсе продавца (эмуляция действия в браузере)"""
        url = "https://a.wb.ru/e/Supplier_Analytics_PeriodFilter_Apl"
        params = {
            "t": "Лента заказов",
            "u": "https://seller.wildberries.ru/content-analytics/order-feed",
            "cid": "7",
            "s": "1280x720x32",
            "w": "544x1456",
            "user_id": "2674786511756018155",  # статичный, можно оставить
            "vbn": "318",
            "nt": "4G",
            "timestamp": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.") + "000Z",
            "timezone_offset": "180",
            "timezone": "Europe/Moscow",
        }
        body = {
            "cp": {
                "reportName": "OrderFeed",
                "periodStart": start_date,
                "periodEnd": end_date,
                "tariffid": 1003,
                "splits": '{"2525":{"expID":"2525","group":"control","isNewAbVersion":false,"decision":true},"2731":{"expID":"2731","group":"control","isNewAbVersion":false,"decision":true},"3133":{"expID":"3133","group":"control","isNewAbVersion":false,"decision":true},"3162":{"expID":"3162","group":"control","isNewSettingsEnabled":false,"decision":true},"3588":{"expID":"3588","group":"control","isShowStories":false,"decision":true},"3871":{"expID":"3871","group":"control","isSupplierCardRedesignV2Exp":false,"decision":true},"3880":{"expID":"3880","group":"control","decision":true},"3928":{"expID":"3928","group":"control","isNewAbVersion":false,"decision":true},"3981":{"expID":"3981","group":"control","isNewAbVersion":false,"decision":true},"4084":{"expID":"4084","experiment":false,"group":"control","decision":true},"4571":{"expID":"4571","group":"control","isNewAbVersion":false,"decision":true},"4906":{"expID":"4906","experiment":true,"group":"test","decision":true},"5278":{"expID":"5278","group":"test","isArticleSubstitutionAb":true,"decision":true},"5375":{"expID":"5375","group":"test","hasOverlayImage":true,"decision":true},"5376":{"expID":"5376","group":"control","isShowRankedAllPagesBanners":false,"decision":true},"5496":{"expID":"5496","experiment":false,"group":"control","decision":true},"5655":{"expID":"5655","group":"control","isPaidServicesAb":false,"decision":true},"5914":{"expID":"5914","group":"test","isNewAbVersion":true,"decision":true},"9990":{"expID":"9990","group":"control","decision":true},"9996":{"expID":"9996","group":"control","isNewAbVersion":false,"decision":true},"9998":{"expID":"9998","group":"test","isNewAbVersion":true,"decision":true}}',
                "currentPageUrl": "https://seller.wildberries.ru/content-analytics/order-feed",
                "uiRootVersion": "v1.91.1",
                "uiRootBuildTime": 1778080582642,
                "isNewAppVersion": 0,
                "language": "ru",
                "idSupplier": "8bcba44c-4a88-5509-b427-ca41a335ff1a",
                "idUser": 96997814,
            },
            "user_ids": {"wba_fp": "833999517f8523b67e7bafb7c2de237c"},
        }
        headers = {
            **self.base_headers,
            "Content-Type": "text/plain"
        }
        response = requests.post(url, params=params, headers=headers, data=json.dumps(body), timeout=30)
        if response.status_code != 200:
            raise Exception(f"Ошибка установки периода: {response.status_code} {response.text}")
        log_event('INFO', 'OrderFeedPrivateAPI.set_period', f'Период {start_date} – {end_date} установлен')

    def fetch_and_save_csv(self, start_date: str, end_date: str, csv_path: str) -> bool:
        """
        Получить отчёт за период и сохранить как CSV.
        Возвращает True при успехе, иначе False.
        """
        try:
            self.set_period(start_date, end_date)

            report_name = f"OrderFeed_{start_date}_{end_date}_{int(time.time())}"
            report_id = self.create_report(report_name, start_date, end_date)

            self.wait_for_done(report_id, timeout=300)

            token = self.get_download_token()

            parquet_bytes = self.download_and_convert_to_parquet(report_id, token)

            df = pd.read_parquet(BytesIO(parquet_bytes))

            os.makedirs(os.path.dirname(csv_path), exist_ok=True)
            df.to_csv(csv_path, index=False, encoding='utf-8-sig')

            log_event('INFO', 'OrderFeedPrivateAPI.fetch_and_save_csv',
                      f'CSV сохранён: {csv_path} (записей: {len(df)})')
            return True

        except Exception as e:
            log_event('ERROR', 'OrderFeedPrivateAPI.fetch_and_save_csv',
                      f'Ошибка: {str(e)}', {'traceback': traceback.format_exc()})
            return False

    def create_report(self, report_name: str, start_date: str, end_date: str) -> str:
        url = "https://seller-content.wildberries.ru/ns/analytics-api/content-analytics/api/v1/file-manager/download"
        report_id = str(uuid.uuid4())
        payload = {
            "id": report_id,
            "userReportName": report_name,
            "reportType": "ORDER_FEED_ORDERS_REPORT",
            "params": {
                "brandNames": [],
                "subjectIDs": [],
                "tagIDs": [],
                "nmIDs": [],
                "currentPeriod": {"start": start_date, "end": end_date},
                "orderBy": {"field": "order.updated", "mode": "desc"},
                "isDefective": False,
                "timezone": "Europe/Moscow"
            }
        }
        headers = {
            **self.base_headers,
            "AuthorizeV3": self.authorize_v3,
            "Wb-Seller-Lk": self.wb_seller_lk,
            "Root-Version": "v1.91.1",
            "Content-Type": "application/json",
        }
        response = requests.post(url, headers=headers, json=payload, timeout=30)
        if response.status_code != 200:
            raise Exception(f"Ошибка создания отчёта: {response.status_code} {response.text}")
        log_event('INFO', 'OrderFeedPrivateAPI.create_report', f'Отчёт создан, ID: {report_id}')
        return report_id

    def wait_for_done(self, report_id: str, timeout: int = 300) -> bool:
        url = "https://seller-content.wildberries.ru/ns/analytics-api/content-analytics/api/v1/file-manager/downloads"
        params = {"report_types": "ORDER_FEED_ORDERS_REPORT"}
        headers = {
            **self.base_headers,
            "AuthorizeV3": self.authorize_v3,
            "Wb-Seller-Lk": self.wb_seller_lk,
            "Root-Version": "v1.91.1",
        }
        start_time = time.time()
        while time.time() - start_time < timeout:
            response = requests.get(url, params=params, headers=headers, timeout=30)
            if response.status_code == 200:
                data = response.json()
                downloads = data.get("data", {}).get("downloads", [])
                for item in downloads:
                    if item.get("id") == report_id:
                        status = item.get("status")
                        log_event('DEBUG', 'OrderFeedPrivateAPI.wait_for_done', f'Статус отчёта: {status}')
                        if status in ("DONE", "SUCCESS"):
                            log_event('INFO', 'OrderFeedPrivateAPI.wait_for_done', f'Отчёт готов')
                            return True
                        elif status in ("FAILED", "CANCELLED"):
                            raise Exception(f"Отчёт провалился со статусом {status}")
            time.sleep(1)
        raise TimeoutError("Превышено время ожидания готовности отчёта")

    def get_download_token(self) -> str:
        url = "https://seller-content.wildberries.ru/ns/suppliers-auth-tokens/suppliers-portal-core/api/v1/tokensjrpc"
        payload = {
            "method": "generateToken",
            "params": {"team": "content-analytics"},
            "jsonrpc": "2.0",
            "id": "json-rpc_56"
        }
        headers = {
            **self.base_headers,
            "AuthorizeV3": self.authorize_v3,
            "Wb-Seller-Lk": self.wb_seller_lk,
            "Root-Version": "v1.91.1",
            "Content-Type": "application/json",
        }
        response = requests.post(url, headers=headers, json=payload, timeout=30)
        if response.status_code != 200:
            raise Exception(f"Ошибка получения токена: {response.status_code} {response.text}")
        result = response.json()
        token = result.get("result", {}).get("token")
        if not token and isinstance(result.get("result"), str):
            token = result["result"]
        if not token:
            raise Exception(f"Токен не найден в ответе: {json.dumps(result, indent=2, ensure_ascii=False)}")
        log_event('INFO', 'OrderFeedPrivateAPI.get_download_token', 'Токен получен')
        return token

    def download_and_convert_to_parquet(self, report_id: str, token: str) -> bytes:
        url = f"https://downloads-content-analytics.wildberries.ru/api/v1/file-manager/download/{report_id}"
        headers = {**self.base_headers, "X-Download-Token": token}
        response = requests.get(url, headers=headers, timeout=60)
        if response.status_code != 200:
            raise Exception(f"Ошибка скачивания: {response.status_code} {response.text}")

        with zipfile.ZipFile(BytesIO(response.content)) as z:
            xlsx_name = next((name for name in z.namelist() if name.lower().endswith('.xlsx')), None)
            if not xlsx_name:
                raise Exception("В архиве не найден XLSX-файл")
            log_event('INFO', 'OrderFeedPrivateAPI.download_and_convert_to_parquet', f'Найден файл: {xlsx_name}')

            with z.open(xlsx_name) as xlsx_file:
                wb = openpyxl.load_workbook(xlsx_file, data_only=True)
                if "Все заказы" not in wb.sheetnames:
                    raise Exception("Лист 'Все заказы' не найден в книге")

                ws = wb["Все заказы"]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) <= 1:
                    raise Exception("Лист 'Все заказы' не содержит данных (только заголовок)")

                header_row = rows[1]  # вторая строка файла (индекс 1)
                data_rows = rows[2:]  # начиная с третьей строки

                header_row = [str(cell) if cell is not None else '' for cell in header_row]
                clean_headers = []
                counter = {}
                for col in header_row:
                    if not col or col == '':
                        col = 'empty'
                    if col in counter:
                        counter[col] += 1
                        col = f"{col}_{counter[col]}"
                    else:
                        counter[col] = 1
                    clean_headers.append(col)

                data = []
                for row in data_rows:
                    row_vals = list(row)[:len(clean_headers)]
                    row_vals = [v if v is not None else '' for v in row_vals]
                    data.append(row_vals)

                df = pd.DataFrame(data, columns=clean_headers)
                buffer = BytesIO()
                df.to_parquet(buffer, index=False, compression='snappy')
                buffer.seek(0)
                log_event('INFO', 'OrderFeedPrivateAPI.download_and_convert_to_parquet',
                          f'Parquet сформирован, строк: {len(df)}')
                return buffer.getvalue()